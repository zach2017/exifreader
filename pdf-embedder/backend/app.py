import base64
import io
import os
import time
import uuid
import mimetypes
from typing import Dict, Any, List, Tuple

from flask import Flask, request, send_file, send_from_directory, jsonify
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from PIL import Image, ExifTags

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

app = Flask(__name__, static_folder="static", static_url_path="")

MAX_UPLOAD_MB = int(os.getenv("MAX_UPLOAD_MB", "50"))
MAX_BYTES = MAX_UPLOAD_MB * 1024 * 1024

# token -> {"ts": epoch_seconds, "pdf_bytes": bytes, "attachments": {name: bytes}, "images": [{"id":..., "bytes":..., "mime":..., "page":...}]}
STORE: Dict[str, Dict[str, Any]] = {}
STORE_TTL_SECONDS = int(os.getenv("STORE_TTL_SECONDS", "900"))  # 15 minutes default


def _too_large(content_length):
    return content_length is not None and content_length > MAX_BYTES


def _cleanup_store():
    now = time.time()
    expired = [k for k, v in STORE.items() if (now - v.get("ts", now)) > STORE_TTL_SECONDS]
    for k in expired:
        STORE.pop(k, None)


def _guess_mime(filename: str) -> str:
    mt, _ = mimetypes.guess_type(filename)
    return mt or "application/octet-stream"


def make_image_overlay_pdf(image_bytes, page_w, page_h, placement="top-right"):
    img = Image.open(io.BytesIO(image_bytes)).convert("RGBA")

    target_w = page_w * 0.30
    scale = target_w / img.width
    target_h = img.height * scale

    if target_h > page_h * 0.35:
        scale = (page_h * 0.35) / img.height
        target_w = img.width * scale
        target_h = img.height * scale

    margin = 18
    if placement == "top-left":
        x, y = margin, page_h - margin - target_h
    elif placement == "bottom-left":
        x, y = margin, margin
    elif placement == "bottom-right":
        x, y = page_w - margin - target_w, margin
    else:
        x, y = page_w - margin - target_w, page_h - margin - target_h

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(page_w, page_h))
    c.drawImage(ImageReader(img), x, y, width=target_w, height=target_h, mask="auto")
    c.showPage()
    c.save()
    return buf.getvalue()


def _read_attachments(reader: PdfReader) -> Dict[str, bytes]:
    """
    Robust extraction of embedded file attachments ("File Attachments") via NameTree:
    /Root -> /Names -> /EmbeddedFiles -> /Names
    Returns: {filename: bytes}
    """
    attachments: Dict[str, bytes] = {}

    def _safe_text(obj) -> str:
        try:
            return str(obj)
        except Exception:
            return "attachment.bin"

    try:
        root = reader.trailer.get("/Root")
        if not root:
            return attachments

        names = root.get("/Names")
        if not names:
            return attachments

        embedded = names.get("/EmbeddedFiles")
        if not embedded:
            return attachments

        names_array = embedded.get("/Names")
        if not names_array:
            return attachments

        for i in range(0, len(names_array), 2):
            try:
                fname_obj = names_array[i]
                filespec = names_array[i + 1]
            except Exception:
                continue

            fname = _safe_text(fname_obj)

            try:
                ef = filespec.get("/EF") or {}
                # Streams can be stored under several keys; prefer /F then fall back.
                for key in ("/F", "/UF", "/DOS", "/Mac", "/Unix"):
                    st = ef.get(key)
                    if st:
                        attachments[fname] = st.get_data()
                        break
            except Exception:
                pass

    except Exception:
        pass

    # Last resort: pypdf's helper (varies by version/pdf structure)
    if not attachments:
        try:
            att = getattr(reader, "attachments", None)
            if att and isinstance(att, dict):
                for k, v in att.items():
                    if isinstance(v, (bytes, bytearray)):
                        attachments[str(k)] = bytes(v)
                    else:
                        try:
                            attachments[str(k)] = bytes(v[0])
                        except Exception:
                            pass
        except Exception:
            pass

    return attachments


def _xlsx_preview(xlsx_bytes: bytes, max_rows: int = 50, max_cols: int = 20) -> Dict[str, Any]:
    if load_workbook is None:
        return {"type": "error", "message": "openpyxl not available in server image."}

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        rows.append([("" if v is None else v) for v in row[:max_cols]])
    wb.close()

    def norm(v):
        if isinstance(v, (int, float)):
            return v
        return str(v)

    normalized = [[norm(v) for v in r] for r in rows]
    return {"type": "table", "sheet": ws.title, "rows": normalized}


def _csv_preview(csv_bytes: bytes, max_rows: int = 80) -> Dict[str, Any]:
    try:
        text = csv_bytes.decode("utf-8", errors="replace")
        lines = text.splitlines()
        rows = [l.split(",") for l in lines[:max_rows]]
        return {"type": "table", "sheet": "CSV", "rows": rows}
    except Exception as e:
        return {"type": "error", "message": f"CSV preview failed: {e}"}


def _exif_to_dict(img: Image.Image) -> Dict[str, Any]:
    """
    Extract EXIF (if present) into a JSON-serializable dict.
    Note: Many PDF-extracted images won't contain EXIF because they may be re-encoded.
    """
    out: Dict[str, Any] = {}
    try:
        exif = img.getexif()
        if not exif:
            return out
        tag_map = ExifTags.TAGS
        for tag_id, value in exif.items():
            name = tag_map.get(tag_id, str(tag_id))
            # Make sure it's JSON-friendly
            if isinstance(value, bytes):
                try:
                    value = value.decode("utf-8", errors="replace")
                except Exception:
                    value = base64.b64encode(value).decode("ascii")
            out[name] = value
    except Exception:
        pass
    return out


def _extract_images_from_pdf(reader: PdfReader, max_images: int = 50) -> List[Dict[str, Any]]:
    """
    Extract images from PDF pages by scanning XObject images.
    Produces:
      - original_bytes + original_mime (best guess)
      - preview_bytes as PNG when we can reconstruct (for reliable browser preview)
    """
    images: List[Dict[str, Any]] = []
    seen = 0

    def _filters(o):
        f = o.get("/Filter")
        if f is None:
            return []
        if isinstance(f, list):
            return [str(x) for x in f]
        return [str(f)]

    def _colorspace(o):
        cs = o.get("/ColorSpace")
        if cs is None:
            return "/DeviceRGB"
        try:
            if isinstance(cs, list) and cs:
                return str(cs[0])
            return str(cs)
        except Exception:
            return "/DeviceRGB"

    def _pil_mode_from_cs(cs_name: str) -> str:
        if cs_name == "/DeviceGray":
            return "L"
        if cs_name == "/DeviceCMYK":
            return "CMYK"
        return "RGB"

    for page_index, page in enumerate(reader.pages):
        if seen >= max_images:
            break
        try:
            resources = page.get("/Resources") or {}
            xobj = resources.get("/XObject")
            if not xobj:
                continue

            for name, obj in xobj.items():
                if seen >= max_images:
                    break
                try:
                    o = obj.get_object()
                    if o.get("/Subtype") != "/Image":
                        continue

                    filters = _filters(o)
                    w = int(o.get("/Width") or 0)
                    h = int(o.get("/Height") or 0)
                    bpc = int(o.get("/BitsPerComponent") or 8)
                    cs_name = _colorspace(o)
                    mode = _pil_mode_from_cs(cs_name)

                    data = o.get_data()

                    original_mime = "application/octet-stream"
                    original_ext = "bin"
                    if "/DCTDecode" in filters:
                        original_mime, original_ext = "image/jpeg", "jpg"
                    elif "/JPXDecode" in filters:
                        original_mime, original_ext = "image/jp2", "jp2"
                    elif "/CCITTFaxDecode" in filters:
                        original_mime, original_ext = "image/tiff", "tif"

                    img_id = f"p{page_index+1}_{seen+1}_{str(name).strip('/')}"
                    rec: Dict[str, Any] = {
                        "id": img_id,
                        "page": page_index + 1,
                        "name": f"{img_id}.{original_ext}",
                        "original_mime": original_mime,
                        "original_bytes": data,
                        "preview_mime": None,
                        "preview_bytes": None,
                    }

                    # JPEG: browser-friendly as-is
                    if original_mime == "image/jpeg":
                        rec["preview_mime"] = "image/jpeg"
                        rec["preview_bytes"] = data

                    # FlateDecode or no filter: decoded raw pixels -> reconstruct
                    elif ("/FlateDecode" in filters) or (len(filters) == 0):
                        try:
                            if w > 0 and h > 0 and bpc in (1, 8):
                                if bpc == 1:
                                    img = Image.frombytes("1", (w, h), data).convert("L")
                                else:
                                    channels = 1 if mode == "L" else (4 if mode == "CMYK" else 3)
                                    expected = w * h * channels
                                    if len(data) >= expected:
                                        img = Image.frombytes(mode, (w, h), data[:expected])
                                        if mode == "CMYK":
                                            img = img.convert("RGB")
                                    else:
                                        img = None

                                if img is not None:
                                    buf = io.BytesIO()
                                    img.save(buf, format="PNG")
                                    rec["preview_mime"] = "image/png"
                                    rec["preview_bytes"] = buf.getvalue()
                        except Exception:
                            pass

                    # JPXDecode: try convert to PNG if Pillow supports; otherwise no preview
                    elif original_mime == "image/jp2":
                        try:
                            img = Image.open(io.BytesIO(data))
                            buf = io.BytesIO()
                            img.save(buf, format="PNG")
                            rec["preview_mime"] = "image/png"
                            rec["preview_bytes"] = buf.getvalue()
                        except Exception:
                            rec["preview_mime"] = None
                            rec["preview_bytes"] = None

                    # Anything else: attempt Pillow open -> PNG preview
                    else:
                        try:
                            img = Image.open(io.BytesIO(data))
                            buf = io.BytesIO()
                            img.save(buf, format="PNG")
                            rec["preview_mime"] = "image/png"
                            rec["preview_bytes"] = buf.getvalue()
                        except Exception:
                            pass

                    # If we only have a PNG preview and original is opaque, set download to PNG
                    if rec["preview_mime"] == "image/png" and rec["original_mime"] == "application/octet-stream":
                        rec["name"] = f"{img_id}.png"
                        rec["original_mime"] = "image/png"
                        rec["original_bytes"] = rec["preview_bytes"] or rec["original_bytes"]

                    images.append(rec)
                    seen += 1

                except Exception:
                    continue
        except Exception:
            continue

    return images


def _pdf_info(reader: PdfReader) -> Dict[str, Any]:
    info = {}
    try:
        md = reader.metadata
        if md:
            for k, v in md.items():
                info[str(k)] = "" if v is None else str(v)
    except Exception:
        pass
    return info


@app.get("/")
def index():
    return send_from_directory(app.static_folder, "index.html")


@app.post("/api/embed")
def embed():
    if _too_large(request.content_length):
        return jsonify({"error": f"Upload too large (max {MAX_UPLOAD_MB} MB)."}), 413

    if "pdf" not in request.files or "extra" not in request.files:
        return jsonify({"error": "Missing required files: pdf and extra"}), 400

    pdf_file = request.files["pdf"]
    extra_file = request.files["extra"]

    pdf_bytes = pdf_file.read()
    extra_bytes = extra_file.read()
    extra_name = extra_file.filename or "attachment.bin"

    if not pdf_bytes:
        return jsonify({"error": "PDF is empty"}), 400
    if not extra_bytes:
        return jsonify({"error": "Extra file is empty"}), 400

    reader = PdfReader(io.BytesIO(pdf_bytes))
    writer = PdfWriter()

    for p in reader.pages:
        writer.add_page(p)

    # Embed attachment
    try:
        writer.add_attachment(extra_name, extra_bytes)
    except Exception as e:
        return jsonify({"error": f"Failed to embed attachment: {e}"}), 500

    # If image, also stamp onto first page
    name_lower = extra_name.lower()
    is_image = name_lower.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif")) or (extra_file.mimetype or "").startswith("image/")
    if is_image and len(writer.pages) > 0:
        placement = request.form.get("placement", "top-right")
        first = writer.pages[0]
        w, h = float(first.mediabox.width), float(first.mediabox.height)
        try:
            overlay_pdf = make_image_overlay_pdf(extra_bytes, w, h, placement=placement)
            overlay_reader = PdfReader(io.BytesIO(overlay_pdf))
            first.merge_page(overlay_reader.pages[0])
        except Exception as e:
            return jsonify({"error": f"Embedded attachment but failed to stamp image: {e}"}), 500

    out = io.BytesIO()
    writer.write(out)
    out.seek(0)

    return send_file(out, as_attachment=True, download_name="updated.pdf", mimetype="application/pdf")


@app.post("/api/verify")
def verify():
    _cleanup_store()

    if _too_large(request.content_length):
        return jsonify({"error": f"Upload too large (max {MAX_UPLOAD_MB} MB)."}), 413

    if "pdf" not in request.files:
        return jsonify({"error": "Missing required file: pdf"}), 400

    pdf_bytes = request.files["pdf"].read()
    if not pdf_bytes:
        return jsonify({"error": "PDF is empty"}), 400

    reader = PdfReader(io.BytesIO(pdf_bytes))
    attachments = _read_attachments(reader)

    token = str(uuid.uuid4())
    STORE[token] = {"ts": time.time(), "pdf_bytes": pdf_bytes, "attachments": attachments, "images": []}

    items = []
    for name, b in attachments.items():
        items.append({"name": name, "size": len(b), "mime": _guess_mime(name)})

    return jsonify({
        "token": token,
        "attachment_count": len(items),
        "attachments": sorted(items, key=lambda x: x["name"].lower())
    })


@app.get("/api/verify/attachment")
def verify_attachment():
    _cleanup_store()
    token = request.args.get("token", "")
    name = request.args.get("name", "")

    if not token or token not in STORE:
        return jsonify({"error": "Invalid/expired token. Re-upload PDF to verify again."}), 400

    attachments = STORE[token].get("attachments", {})
    if name not in attachments:
        return jsonify({"error": "Attachment not found in this PDF."}), 404

    data = attachments[name]
    mt = _guess_mime(name)
    return send_file(io.BytesIO(data), as_attachment=True, download_name=name, mimetype=mt)


@app.get("/api/verify/preview")
def verify_preview():
    _cleanup_store()
    token = request.args.get("token", "")
    name = request.args.get("name", "")

    if not token or token not in STORE:
        return jsonify({"type": "error", "message": "Invalid/expired token. Re-upload PDF to verify again."}), 400

    attachments = STORE[token].get("attachments", {})
    if name not in attachments:
        return jsonify({"type": "error", "message": "Attachment not found in this PDF."}), 404

    data = attachments[name]
    lower = name.lower()
    mime = _guess_mime(name)

    if lower.endswith((".png", ".jpg", ".jpeg", ".webp", ".gif")) or mime.startswith("image/"):
        b64 = base64.b64encode(data).decode("ascii")
        return jsonify({"type": "image", "mime": mime, "dataUrl": f"data:{mime};base64,{b64}"})

    if lower.endswith((".xlsx", ".xlsm")):
        return jsonify(_xlsx_preview(data))

    if lower.endswith(".csv"):
        return jsonify(_csv_preview(data))

    return jsonify({
        "type": "info",
        "name": name,
        "size": len(data),
        "mime": mime,
        "message": "No built-in preview for this file type. You can still download it."
    })


@app.post("/api/extract")
def extract():
    """
    Upload a PDF, extract:
      - PDF metadata (Info dictionary / XMP as available via pypdf)
      - Embedded file attachments (EmbeddedFiles NameTree)
      - Inline/page images (XObject images)
    Returns a token to download/preview extracted images and attachments.
    """
    _cleanup_store()

    if _too_large(request.content_length):
        return jsonify({"error": f"Upload too large (max {MAX_UPLOAD_MB} MB)."}), 413

    if "pdf" not in request.files:
        return jsonify({"error": "Missing required file: pdf"}), 400

    pdf_bytes = request.files["pdf"].read()
    if not pdf_bytes:
        return jsonify({"error": "PDF is empty"}), 400

    reader = PdfReader(io.BytesIO(pdf_bytes))
    info = _pdf_info(reader)

    attachments = _read_attachments(reader)
    images = _extract_images_from_pdf(reader)

    token = str(uuid.uuid4())
    STORE[token] = {"ts": time.time(), "pdf_bytes": pdf_bytes, "attachments": attachments, "images": images}

    att_list = [{"name": n, "size": len(b), "mime": _guess_mime(n)} for n, b in attachments.items()]
    img_list = [{"id": im["id"], "name": im["name"], "page": im["page"], "mime": im.get("original_mime"), "size": len(im.get("original_bytes") or b"")} for im in images]

    return jsonify({
        "token": token,
        "pdf_metadata": info,
        "attachment_count": len(att_list),
        "attachments": sorted(att_list, key=lambda x: x["name"].lower()),
        "image_count": len(img_list),
        "images": img_list
    })


@app.get("/api/extract/image")
def extract_image_download():
    _cleanup_store()
    token = request.args.get("token", "")
    image_id = request.args.get("id", "")

    if not token or token not in STORE:
        return jsonify({"error": "Invalid/expired token. Re-upload PDF to extract again."}), 400

    images = STORE[token].get("images", [])
    hit = next((im for im in images if im.get("id") == image_id), None)
    if not hit:
        return jsonify({"error": "Image not found."}), 404

    return send_file(io.BytesIO(hit.get("original_bytes") or b""), as_attachment=True, download_name=hit["name"], mimetype=(hit.get("original_mime") or "application/octet-stream"))


@app.get("/api/extract/image_preview")
def extract_image_preview():
    _cleanup_store()
    token = request.args.get("token", "")
    image_id = request.args.get("id", "")

    if not token or token not in STORE:
        return jsonify({"type": "error", "message": "Invalid/expired token. Re-upload PDF to extract again."}), 400

    images = STORE[token].get("images", [])
    hit = next((im for im in images if im.get("id") == image_id), None)
    if not hit:
        return jsonify({"type": "error", "message": "Image not found."}), 404

    data = hit.get("original_bytes") or b""
    mime = hit.get("original_mime") or "application/octet-stream"

    # Prefer prebuilt preview bytes (PNG/JPEG) for reliable browser display
    pbytes = hit.get("preview_bytes")
    pmime = hit.get("preview_mime")

    exif = {}
    data_url = None

    try:
        # EXIF is only meaningful for formats Pillow understands (usually JPEG)
        img = Image.open(io.BytesIO(data))
        exif = _exif_to_dict(img)
    except Exception:
        exif = {}

    if pbytes and pmime:
        b64 = base64.b64encode(pbytes).decode("ascii")
        data_url = f"data:{pmime};base64,{b64}"
    else:
        # No safe preview
        data_url = None

    return jsonify({
        "type": "image",
        "id": image_id,
        "name": hit["name"],
        "page": hit["page"],
        "mime": mime,
        "exif": exif,
        "dataUrl": data_url
    })


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "8080")))